#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Unit Tests for Excel Editor New Features (v1.3.0)
Testing: add_sheet, delete_sheet, set_cell_format, set_formula
"""

import unittest
import os
import tempfile
import shutil
from pathlib import Path

import sys
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from src.excel_editor import ExcelEditor
from openpyxl import Workbook, load_workbook


class TestExcelEditorNewFeatures(unittest.TestCase):
    """測試 Excel Editor v1.3.0 新功能"""
    
    @classmethod
    def setUpClass(cls):
        """設置測試環境"""
        cls.test_dir = tempfile.mkdtemp()
        cls.test_xlsx = os.path.join(cls.test_dir, "test.xlsx")
        
        # 創建測試工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws['A1'] = "測試數據"
        ws['B1'] = 100
        ws['C1'] = 200
        wb.save(cls.test_xlsx)
    
    @classmethod
    def tearDownClass(cls):
        """清理測試環境"""
        shutil.rmtree(cls.test_dir)
    
    def setUp(self):
        """每個測試前的準備"""
        self.output_file = os.path.join(self.test_dir, f"output_{self._testMethodName}.xlsx")
        shutil.copy(self.test_xlsx, self.output_file)
        self.editor = ExcelEditor(self.output_file)
    
    def test_add_sheet_basic(self):
        """測試新增工作表"""
        result = self.editor.add_sheet("NewSheet")
        self.assertTrue(result)
        
        self.editor.save()
        wb = load_workbook(self.output_file)
        self.assertIn("NewSheet", wb.sheetnames)
    
    def test_add_sheet_with_position(self):
        """測試在特定位置新增工作表"""
        result = self.editor.add_sheet("FirstSheet", position=0)
        self.assertTrue(result)
        
        self.editor.save()
        wb = load_workbook(self.output_file)
        self.assertEqual(wb.sheetnames[0], "FirstSheet")
    
    def test_add_sheet_duplicate_name(self):
        """測試新增重複名稱的工作表"""
        self.editor.add_sheet("TestSheet")
        result = self.editor.add_sheet("TestSheet")
        self.assertFalse(result)
    
    def test_delete_sheet_basic(self):
        """測試刪除工作表"""
        # 先新增一個工作表
        self.editor.add_sheet("ToDelete")
        self.editor.save()
        
        # 重新載入並刪除
        editor2 = ExcelEditor(self.output_file)
        result = editor2.delete_sheet("ToDelete")
        self.assertTrue(result)
        
        editor2.save()
        wb = load_workbook(self.output_file)
        self.assertNotIn("ToDelete", wb.sheetnames)
    
    def test_delete_sheet_last_one(self):
        """測試刪除唯一的工作表"""
        result = self.editor.delete_sheet("Sheet1")
        self.assertFalse(result)
    
    def test_delete_sheet_not_exist(self):
        """測試刪除不存在的工作表"""
        result = self.editor.delete_sheet("NonExistent")
        self.assertFalse(result)
    
    def test_set_cell_format_basic(self):
        """測試設定儲存格格式"""
        result = self.editor.set_cell_format(
            "Sheet1",
            "A1",
            bold=True,
            font_size=14
        )
        self.assertTrue(result)
        
        self.editor.save()
        wb = load_workbook(self.output_file)
        cell = wb['Sheet1']['A1']
        self.assertTrue(cell.font.bold)
        self.assertEqual(cell.font.size, 14)
    
    def test_set_cell_format_with_color(self):
        """測試設定背景顏色"""
        result = self.editor.set_cell_format(
            "Sheet1",
            "B1",
            bg_color="FFFF00"  # 黃色
        )
        self.assertTrue(result)
        
        self.editor.save()
        wb = load_workbook(self.output_file)
        cell = wb['Sheet1']['B1']
        self.assertEqual(cell.fill.start_color.rgb, "FFFF00")
    
    def test_set_cell_format_with_alignment(self):
        """測試設定對齊方式"""
        result = self.editor.set_cell_format(
            "Sheet1",
            "C1",
            alignment="center"
        )
        self.assertTrue(result)
        
        self.editor.save()
        wb = load_workbook(self.output_file)
        cell = wb['Sheet1']['C1']
        self.assertEqual(cell.alignment.horizontal, "center")
    
    def test_set_cell_format_invalid_sheet(self):
        """測試無效的工作表名稱"""
        result = self.editor.set_cell_format(
            "InvalidSheet",
            "A1",
            bold=True
        )
        self.assertFalse(result)
    
    def test_set_formula_sum(self):
        """測試設定 SUM 公式"""
        # 先填充一些數據
        self.editor.update_cell("Sheet1", "A2", 10)
        self.editor.update_cell("Sheet1", "A3", 20)
        self.editor.update_cell("Sheet1", "A4", 30)
        
        result = self.editor.set_formula("Sheet1", "A5", "=SUM(A2:A4)")
        self.assertTrue(result)
        
        self.editor.save()
        wb = load_workbook(self.output_file)
        cell = wb['Sheet1']['A5']
        self.assertEqual(cell.value, "=SUM(A2:A4)")
    
    def test_set_formula_average(self):
        """測試設定 AVERAGE 公式"""
        result = self.editor.set_formula("Sheet1", "D1", "=AVERAGE(B1:C1)")
        self.assertTrue(result)
        
        self.editor.save()
        wb = load_workbook(self.output_file)
        cell = wb['Sheet1']['D1']
        self.assertEqual(cell.value, "=AVERAGE(B1:C1)")
    
    def test_set_formula_invalid_sheet(self):
        """測試在無效工作表設定公式"""
        result = self.editor.set_formula("InvalidSheet", "A1", "=1+1")
        self.assertFalse(result)


class TestExcelEditorIntegration(unittest.TestCase):
    """整合測試：組合多個功能"""
    
    def setUp(self):
        """準備測試環境"""
        self.test_dir = tempfile.mkdtemp()
        self.test_file = os.path.join(self.test_dir, "integration_test.xlsx")
    
    def tearDown(self):
        """清理"""
        shutil.rmtree(self.test_dir)
    
    def test_create_formatted_report(self):
        """測試創建格式化報表"""
        # 創建新工作簿
        wb = Workbook()
        wb.save(self.test_file)
        
        editor = ExcelEditor(self.test_file)
        
        # 1. 新增報表工作表
        editor.add_sheet("Q1_Report", position=0)
        
        # 2. 設定標題
        editor.update_cell("Q1_Report", "A1", "Q1 2025 財務報表")
        editor.set_cell_format(
            "Q1_Report", "A1",
            bold=True,
            font_size=16,
            bg_color="4472C4",
            alignment="center"
        )
        
        # 3. 添加數據
        editor.update_cell("Q1_Report", "A2", "項目")
        editor.update_cell("Q1_Report", "B2", "金額")
        
        editor.update_cell("Q1_Report", "A3", "營收")
        editor.update_cell("Q1_Report", "B3", 100000)
        
        editor.update_cell("Q1_Report", "A4", "成本")
        editor.update_cell("Q1_Report", "B4", 60000)
        
        # 4. 設定公式
        editor.update_cell("Q1_Report", "A5", "淨利")
        editor.set_formula("Q1_Report", "B5", "=B3-B4")
        
        # 5. 格式化總計行
        editor.set_cell_format(
            "Q1_Report", "A5",
            bold=True,
            bg_color="D9E1F2"
        )
        
        editor.save()
        
        # 驗證
        wb = load_workbook(self.test_file)
        self.assertIn("Q1_Report", wb.sheetnames)
        self.assertEqual(wb['Q1_Report']['A1'].value, "Q1 2025 財務報表")
        self.assertEqual(wb['Q1_Report']['B5'].value, "=B3-B4")


if __name__ == '__main__':
    unittest.main(verbosity=2)
