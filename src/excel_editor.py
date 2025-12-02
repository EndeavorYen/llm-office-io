#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Interactive Excel Editor
強大的 Excel 互動式編輯工具
支援透過自然語言指令修改 Excel 內容
"""

from typing import Optional, List, Any, Tuple
import os
import sys
import argparse

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False

from .constants import (
    SUCCESS_SYMBOL,
    ERROR_SYMBOL,
    WARNING_SYMBOL,
    MAX_ROWS_DISPLAY,
    MAX_COLS_DISPLAY
)

# Excel 相關常量
EXCEL_EXTENSION = '.xlsx'
DEFAULT_SHEET_NAME = 'Sheet1'


class ExcelEditor:
    """Excel 編輯器類"""
    
    def __init__(self, filepath: str) -> None:
        """初始化 Excel 編輯器
        
        Args:
            filepath: Excel 檔案路徑
            
        Raises:
            FileNotFoundError: 當檔案不存在時
            ValueError: 當檔案格式不支援時
            RuntimeError: 當無法開啟 Excel 檔案時
        """
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"檔案不存在: {filepath}")
        
        if not filepath.endswith(EXCEL_EXTENSION):
            raise ValueError(f"不支援的檔案格式，需要 {EXCEL_EXTENSION}: {filepath}")
        
        try:
            self.filepath = filepath
            self.wb = load_workbook(filepath)
        except Exception as e:
            raise RuntimeError(f"無法開啟 Excel 檔案: {e}") from e
    
    def save(self, output_path: Optional[str] = None) -> None:
        """儲存 Excel 檔案
        
        Args:
            output_path: 輸出路徑，None 表示覆蓋原檔案
        """
        save_path = output_path or self.filepath
        try:
            self.wb.save(save_path)
            print(f"{SUCCESS_SYMBOL} Excel 檔案已儲存: {save_path}")
        except Exception as e:
            print(f"{ERROR_SYMBOL} 儲存失敗: {e}")
            raise
    
    def list_sheets(self) -> None:
        """列出所有工作表及基本資訊"""
        print(f"\n=== Excel 檔案結構 (共 {len(self.wb.sheetnames)} 個工作表) ===\n")
        
        for i, sheet_name in enumerate(self.wb.sheetnames, 1):
            ws = self.wb[sheet_name]
            active_mark = " [活動]" if ws == self.wb.active else ""
            print(f"[工作表 {i}] {sheet_name}{active_mark}")
            print(f"  行數: {ws.max_row}, 列數: {ws.max_column}")
            print()
    
    def view_sheet(self, sheet_name: Optional[str] = None, max_rows: int = MAX_ROWS_DISPLAY) -> None:
        """查看工作表內容
        
        Args:
            sheet_name: 工作表名稱，None 表示活動工作表
            max_rows: 最大顯示行數
        """
        if sheet_name:
            if not self._validate_sheet_name(sheet_name):
                return
            ws = self.wb[sheet_name]
        else:
            ws = self.wb.active
            sheet_name = ws.title
        
        print(f"\n=== 工作表: {sheet_name} ===\n")
        
        # 顯示前幾行
        rows_to_show = min(max_rows, ws.max_row)
        cols_to_show = min(MAX_COLS_DISPLAY, ws.max_column)
        
        # 顯示標頭
        headers = []
        for col in range(1, cols_to_show + 1):
            headers.append(get_column_letter(col))
        print("   " + "  ".join(f"{h:>8}" for h in headers))
        print("   " + "-" * (10 * len(headers)))
        
        # 顯示資料
        for row in range(1, rows_to_show + 1):
            row_data = []
            for col in range(1, cols_to_show + 1):
                cell_value = ws.cell(row, col).value
                if cell_value is None:
                    cell_value = ""
                row_data.append(str(cell_value)[:8])
            print(f"{row:2} " + "  ".join(f"{val:>8}" for val in row_data))
        
        if ws.max_row > max_rows:
            print(f"\n... 還有 {ws.max_row - max_rows} 行未顯示")
        if ws.max_column > MAX_COLS_DISPLAY:
            print(f"... 還有 {ws.max_column - MAX_COLS_DISPLAY} 列未顯示")
    
    def replace_text(
        self, 
        old_text: str, 
        new_text: str, 
        sheet_name: Optional[str] = None
    ) -> int:
        """替換文字
        
        Args:
            old_text: 要替換的文字
            new_text: 新文字
            sheet_name: 指定工作表名稱，None 表示所有工作表
            
        Returns:
            int: 實際替換的次數
        """
        if not old_text:
            print(f"{ERROR_SYMBOL} 要替換的文字不能為空")
            return 0
        
        if sheet_name:
            if not self._validate_sheet_name(sheet_name):
                return 0
            sheets_to_process = [self.wb[sheet_name]]
        else:
            sheets_to_process = [self.wb[name] for name in self.wb.sheetnames]
        
        replaced_count = 0
        
        # 使用進度條（如果處理多個工作表）
        iterator = sheets_to_process
        if HAS_TQDM and len(sheets_to_process) > 1:
            iterator = tqdm(sheets_to_process, desc="處理工作表", leave=False)
        
        for ws in iterator:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and old_text in cell.value:
                        cell.value = cell.value.replace(old_text, new_text)
                        replaced_count += 1
        
        if replaced_count > 0:
            scope = f"工作表 {sheet_name}" if sheet_name else "所有工作表"
            print(f"{SUCCESS_SYMBOL} 在{scope}中替換了 {replaced_count} 處「{old_text}」→「{new_text}」")
        else:
            print(f"{ERROR_SYMBOL} 找不到「{old_text}」")
        
        return replaced_count
    
    def update_cell(self, sheet_name: str, cell_ref: str, value: Any) -> bool:
        """更新儲存格值
        
        Args:
            sheet_name: 工作表名稱
            cell_ref: 儲存格參照 (如 A1, B2)
            value: 新值
            
        Returns:
            bool: 是否更新成功
        """
        if not self._validate_sheet_name(sheet_name):
            return False
        
        try:
            ws = self.wb[sheet_name]
            old_value = ws[cell_ref].value
            ws[cell_ref] = value
            print(f"{SUCCESS_SYMBOL} 已更新 {sheet_name}!{cell_ref}")
            print(f"  舊值: {old_value}")
            print(f"  新值: {value}")
            return True
        except Exception as e:
            print(f"{ERROR_SYMBOL} 更新失敗: {e}")
            return False
    
    def add_row(
        self, 
        sheet_name: str, 
        data: List[Any], 
        position: Optional[int] = None
    ) -> bool:
        """新增行
        
        Args:
            sheet_name: 工作表名稱
            data: 行資料列表
            position: 插入位置（行號），None 表示在最後
            
        Returns:
            bool: 是否新增成功
        """
        if not self._validate_sheet_name(sheet_name):
            return False
        
        ws = self.wb[sheet_name]
        
        if position is None:
            # 在最後新增
            row_num = ws.max_row + 1
            for col, value in enumerate(data, 1):
                ws.cell(row_num, col, value)
            print(f"{SUCCESS_SYMBOL} 已在 {sheet_name} 最後新增一行（第 {row_num} 行）")
        else:
            # 插入到指定位置
            ws.insert_rows(position)
            for col, value in enumerate(data, 1):
                ws.cell(position, col, value)
            print(f"{SUCCESS_SYMBOL} 已在 {sheet_name} 第 {position} 行插入資料")
        
        return True
    
    def delete_row(self, sheet_name: str, row_number: int) -> bool:
        """刪除行
        
        Args:
            sheet_name: 工作表名稱
            row_number: 行號
            
        Returns:
            bool: 是否刪除成功
        """
        if not self._validate_sheet_name(sheet_name):
            return False
        
        ws = self.wb[sheet_name]
        
        if row_number < 1 or row_number > ws.max_row:
            print(f"{ERROR_SYMBOL} 行號 {row_number} 超出範圍（1-{ws.max_row}）")
            return False
        
        ws.delete_rows(row_number)
        print(f"{SUCCESS_SYMBOL} 已刪除 {sheet_name} 第 {row_number} 行")
        return True
    
    def find_cells(
        self, 
        search_text: str, 
        sheet_name: Optional[str] = None
    ) -> List[Tuple[str, str, Any]]:
        """搜尋包含特定文字的儲存格
        
        Args:
            search_text: 搜尋文字
            sheet_name: 工作表名稱，None 表示所有工作表
            
        Returns:
            List[Tuple[str, str, Any]]: [(工作表名, 儲存格參照, 值)]
        """
        if not search_text:
            print(f"{ERROR_SYMBOL} 搜尋文字不能為空")
            return []
        
        if sheet_name:
            if not self._validate_sheet_name(sheet_name):
                return []
            sheets_to_search = [(sheet_name, self.wb[sheet_name])]
        else:
            sheets_to_search = [(name, self.wb[name]) for name in self.wb.sheetnames]
        
        results = []
        
        for ws_name, ws in sheets_to_search:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and search_text in cell.value:
                        results.append((ws_name, cell.coordinate, cell.value))
        
        if results:
            print(f"\n{SUCCESS_SYMBOL} 找到 {len(results)} 個符合的儲存格:\n")
            for ws_name, cell_ref, value in results[:20]:  # 只顯示前 20 個
                print(f"  {ws_name}!{cell_ref}: {value}")
            if len(results) > 20:
                print(f"\n  ... 還有 {len(results) - 20} 個結果未顯示")
        else:
            print(f"{ERROR_SYMBOL} 找不到包含「{search_text}」的儲存格")
        
        return results

    def add_sheet(self, sheet_name: str, position: Optional[int] = None) -> bool:
        """新增工作表
        
        Args:
            sheet_name: 工作表名稱
            position: 插入位置，None 表示最後
            
        Returns:
            bool: 是否成功新增
        """
        if sheet_name in self.wb.sheetnames:
            print(f"{ERROR_SYMBOL} 工作表「{sheet_name}」已存在")
            return False
        
        try:
            if position is None:
                self.wb.create_sheet(sheet_name)
            else:
                self.wb.create_sheet(sheet_name, position)
            
            print(f"{SUCCESS_SYMBOL} 已新增工作表: {sheet_name}")
            return True
        except Exception as e:
            print(f"{ERROR_SYMBOL} 新增工作表失敗: {e}")
            return False
    
    def delete_sheet(self, sheet_name: str) -> bool:
        """刪除工作表
        
        Args:
            sheet_name: 工作表名稱
            
        Returns:
            bool: 是否成功刪除
        """
        if not self._validate_sheet_name(sheet_name):
            return False
        
        if len(self.wb.sheetnames) == 1:
            print(f"{ERROR_SYMBOL} 無法刪除唯一的工作表")
            return False
        
        try:
            del self.wb[sheet_name]
            print(f"{SUCCESS_SYMBOL} 已刪除工作表: {sheet_name}")
            return True
        except Exception as e:
            print(f"{ERROR_SYMBOL} 刪除工作表失敗: {e}")
            return False
    
    def set_cell_format(
        self,
        sheet_name: str,
        cell_ref: str,
        bold: bool = False,
        font_size: int = 11,
        bg_color: Optional[str] = None,
        alignment: Optional[str] = None
    ) -> bool:
        """設定儲存格格式
        
        Args:
            sheet_name: 工作表名稱
            cell_ref: 儲存格參照 (如 A1)
            bold: 是否粗體
            font_size: 字體大小
            bg_color: 背景顏色 (16進位，如 'FFFF00' 為黃色)
            alignment: 對齊方式 ('left', 'center', 'right')
            
        Returns:
            bool: 是否成功設定
        """
        if not self._validate_sheet_name(sheet_name):
            return False
        
        try:
            ws = self.wb[sheet_name]
            cell = ws[cell_ref]
            
            # 設定字體
            cell.font = Font(bold=bold, size=font_size)
            
            # 設定背景色
            if bg_color:
                cell.fill = PatternFill(
                    start_color=bg_color,
                    end_color=bg_color,
                    fill_type="solid"
                )
            
            # 設定對齊
            if alignment:
                cell.alignment = Alignment(horizontal=alignment)
            
            print(f"{SUCCESS_SYMBOL} 已設定 {sheet_name}!{cell_ref} 的格式")
            return True
        except Exception as e:
            print(f"{ERROR_SYMBOL} 設定格式失敗: {e}")
            return False
    
    def set_formula(
        self,
        sheet_name: str,
        cell_ref: str,
        formula: str
    ) -> bool:
        """設定儲存格公式
        
        Args:
            sheet_name: 工作表名稱
            cell_ref: 儲存格參照 (如 A1)
            formula: 公式 (如 '=SUM(A1:A10)')
            
        Returns:
            bool: 是否成功設定
        """
        if not self._validate_sheet_name(sheet_name):
            return False
        
        try:
            ws = self.wb[sheet_name]
            ws[cell_ref] = formula
            print(f"{SUCCESS_SYMBOL} 已設定公式: {cell_ref} = {formula}")
            return True
        except Exception as e:
            print(f"{ERROR_SYMBOL} 設定公式失敗: {e}")
            return False
    
    def _validate_sheet_name(self, sheet_name: str) -> bool:
        """驗證工作表名稱是否存在
        
        Args:
            sheet_name: 工作表名稱
            
        Returns:
            bool: 是否存在
        """
        if sheet_name not in self.wb.sheetnames:
            print(f"{ERROR_SYMBOL} 工作表「{sheet_name}」不存在")
            print(f"可用的工作表: {', '.join(self.wb.sheetnames)}")
            return False
        return True


def main() -> None:
    """主函數"""
    parser = argparse.ArgumentParser(
        description='Excel 互動式編輯器',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
範例:
  # 列出所有工作表
  python excel_editor.py data.xlsx list
  
  # 查看工作表
  python excel_editor.py data.xlsx view Sheet1
  
  # 替換文字
  python excel_editor.py data.xlsx replace "舊值" "新值"
  
  # 更新儲存格
  python excel_editor.py data.xlsx update-cell Sheet1 A1 "新值"
  
  # 搜尋儲存格
  python excel_editor.py data.xlsx find "關鍵字"
        '''
    )
    
    parser.add_argument('file', help='Excel 檔案路徑')
    parser.add_argument('--output', '-o', help='輸出檔案路徑（不指定則覆蓋原檔案）')
    
    subparsers = parser.add_subparsers(dest='command', help='編輯命令')
    
    # list: 列出工作表
    subparsers.add_parser('list', help='列出所有工作表')
    
    # view: 查看工作表
    view_parser = subparsers.add_parser('view', help='查看工作表內容')
    view_parser.add_argument('sheet', nargs='?', help='工作表名稱（不指定則為活動工作表）')
    view_parser.add_argument('--max-rows', type=int, default=MAX_ROWS_DISPLAY, help='最大顯示行數')
    
    # replace: 替換文字
    replace_parser = subparsers.add_parser('replace', help='替換文字')
    replace_parser.add_argument('old', help='要替換的文字')
    replace_parser.add_argument('new', help='新文字')
    replace_parser.add_argument('--sheet', help='指定工作表名稱（不指定則全部）')
    
    # update-cell: 更新儲存格
    update_parser = subparsers.add_parser('update-cell', help='更新儲存格值')
    update_parser.add_argument('sheet', help='工作表名稱')
    update_parser.add_argument('cell', help='儲存格參照 (如 A1)')
    update_parser.add_argument('value', help='新值')
    
    # add-row: 新增行
    addrow_parser = subparsers.add_parser('add-row', help='新增行')
    addrow_parser.add_argument('sheet', help='工作表名稱')
    addrow_parser.add_argument('data', nargs='+', help='行資料（多個值）')
    addrow_parser.add_argument('--position', type=int, help='插入位置（不指定則在最後）')
    
    # delete-row: 刪除行
    delrow_parser = subparsers.add_parser('delete-row', help='刪除行')
    delrow_parser.add_argument('sheet', help='工作表名稱')
    delrow_parser.add_argument('row', type=int, help='行號')
    
    # find: 搜尋儲存格
    find_parser = subparsers.add_parser('find', help='搜尋儲存格')
    find_parser.add_argument('text', help='搜尋文字')
    find_parser.add_argument('--sheet', help='指定工作表名稱（不指定則全部）')
    
    args = parser.parse_args()
    
    if not args.command:
        parser.print_help()
        return
    
    # 載入 Excel 檔案
    try:
        editor = ExcelEditor(args.file)
    except (FileNotFoundError, ValueError, RuntimeError) as e:
        print(f"{ERROR_SYMBOL} {e}")
        sys.exit(1)
    
    # 執行命令
    try:
        if args.command == 'list':
            editor.list_sheets()
            return
        
        elif args.command == 'view':
            editor.view_sheet(args.sheet, args.max_rows)
            return
        
        elif args.command == 'replace':
            editor.replace_text(args.old, args.new, args.sheet)
        
        elif args.command == 'update-cell':
            editor.update_cell(args.sheet, args.cell, args.value)
        
        elif args.command == 'add-row':
            editor.add_row(args.sheet, args.data, args.position)
        
        elif args.command == 'delete-row':
            editor.delete_row(args.sheet, args.row)
        
        elif args.command == 'find':
            editor.find_cells(args.text, args.sheet)
            return
        
        # 儲存
        editor.save(args.output)
        
    except Exception as e:
        print(f"{ERROR_SYMBOL} 操作失敗: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
